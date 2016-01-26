#!/usr/bin/env python
# -*- coding: utf-8 -*-

import pickle
from pypif.obj import *
import os
import sys

import csv
from openpyxl import load_workbook

if len(sys.argv) < 2:
    root_dir = os.path.join(os.environ["HOME"], "Example")
else:
    root_dir = sys.argv[1]

def irows(file_name):
    """
    Iterates file rows
    """
    _, ext = os.path.splitext(file_name)
    if ext == '.csv':
        with open(os.path.join(dir_name, file_name),'rU') as table:
            reader = csv.reader(table)
            for row in reader:
                yield row
    else:
        wb = load_workbook(file_name)
        for row in wb.worksheets[0].iter_rows():
            yield [str('' if c.value is None else c.value) for c in row]


def conver_file(file_name):
    data = []
    headings = []
    phases = []
    name, ext = os.path.splitext(file_name)
    output_name = name + ".p"

    for i, row in enumerate(irows(file_name)):
        print i, ":", row
        if i == 0:
            img_no = row[0].lower().replace('table','').replace(' ','')
        elif i == 1:
            caption = row[0]
        elif i == 2:
            for j, column in enumerate(row):
                headings.append(row[j])
        else:
            if row[0] <> '':
                chemical_formula = row[0]
            region = row[1]
            if region == 'A':
                p = 'Proeutectic'
            elif region == 'B':
                p = 'Eutectic'

            phase = AlloyPhase()
            phase.chemical_formula = chemical_formula

            compositions = []
            for k in range(2,7):
                comp = Composition()
                comp.element=headings[k]
                comp.atomic_percent=row[k]
                compositions.append(comp)

            phase.composition = compositions
            phase.table = {'number':img_no, 'caption':caption}
            phase.names = [region,p]

            data.append({'labels':[chemical_formula], 'value':phase})

    pickle.dump(data, open(output_name, 'w'))

for dir_name, _, file_names in os.walk(root_dir):
    for file_name in file_names:
        if file_name.endswith(".csv") or file_name.endswith(".xlsx"):
            print "***", file_name
            conver_file(os.path.join(dir_name, file_name))

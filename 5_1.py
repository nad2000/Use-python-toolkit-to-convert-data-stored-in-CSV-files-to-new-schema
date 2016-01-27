#!/usr/bin/env python
# -*- coding: utf-8 -*-

import pickle
from pypif.obj import *
import os
import sys
import re

import csv
from openpyxl import load_workbook

if len(sys.argv) < 2:
    root_dir = os.path.join(os.environ["HOME"], "Example")
else:
    root_dir = sys.argv[1]

verbose = ('-V' in sys.argv)

def irows(file_name):
    """
    Iterates file rows
    """
    _, ext = os.path.splitext(file_name)
    if ext == '.csv':
        with open(os.path.join(dir_name, file_name),'rU') as table:
            reader = csv.reader(table)
            for row in reader:
                yield row
    else:
        wb = load_workbook(file_name)
        for row in wb.worksheets[0].iter_rows():
            yield [str('' if c.value is None else c.value) for c in row]


def conver_file(file_name):
    data = []
    headings = []
    alloys = []
    name, ext = os.path.splitext(file_name)
    output_name = name + ".p"

    for i, row in enumerate(irows(file_name)):
        if verbose:
            print i, ":", row
        if i == 0:
            table_no = row[0].lower().replace('table','').replace(' ','')
        elif i == 1:
            caption = row[0]
        elif i == 2:
            #for j, column in enumerate(row):
            #    headings.append(row[j])
            headings = [h.split('(')[0].strip() for h in row]
            units = re.findall("\(([%\w]*)\)", ' '.join(row))
            if verbose:
                print "Units:", units
                print "Headings:", headings
        else:
            if row[0] <> '':
                chemical_formula = row[0]

            # Yield strength (MPa) ,
            # Young's modulus E (GPa),
            # Fracture strain (%)                
                
            alloy = Alloy(
                ids = chemical_formula, 
                chemical_formula = chemical_formula) 

            properties = []
            for k in range(1,4):
                prop = Property(name=headings[k])
                prop.scalars = row[k]
                prop.units = units[k-1]
                properties.append(prop)

            alloy.properties = properties
            alloy.table = {'number':table_no, 'caption':caption}

            data.append({'labels':[chemical_formula], 'value':alloy})

    pickle.dump(data, open(output_name, 'w'))

for dir_name, _, file_names in os.walk(root_dir):
    for file_name in file_names:
        #if file_name.endswith(".csv") or file_name.endswith(".xlsx"):
        if file_name in ["5_1.csv"]:
            print "***", file_name
            conver_file(os.path.join(dir_name, file_name))

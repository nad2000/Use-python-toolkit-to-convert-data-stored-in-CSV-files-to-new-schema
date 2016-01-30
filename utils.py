#!/usr/bin/env python
# -*- coding: utf-8 -*-

import pickle
from pypif.obj import *
import os
import sys
import re
from fnmatch import fnmatch

import csv
from openpyxl import load_workbook

if len(sys.argv) < 2:
    root_dir = os.path.join(os.environ["HOME"], "Example")
else:
    root_dir = sys.argv[1]

verbose = ('-V' in sys.argv)


def get_units(row):
    units = []
    for c in row:
        u = re.findall("\(([^)]*)\)", c)
        units.append(u[0] if u else None)
    return units


def irows(file_name, max_columns=None):
    """
    Iterates through file rows
    """
    _, ext = os.path.splitext(file_name)
    if ext == '.csv':
        with open(file_name, "rU") as table:
            reader = csv.reader(table)
            for i, row in enumerate(reader):
                if max_columns is not None:
                    row = row[:max_columns]
                row = [v.strip() for v in row]
                if verbose and not all([c == '' for c in row]):
                    print i, ":", row
                yield row
    else:
        wb = load_workbook(file_name)
        for i, ws_row in enumerate(wb.worksheets[0].iter_rows()):
            row = [str('' if c.value is None else c.value).strip() for c in ws_row]
            if max_columns is not None:
                row = row[:max_columns]
            if verbose:
                print i, ":", row
            yield row


def ifiles(files=None, root_dir=root_dir):
    """
    Iterates through all files in the directroy
    and its subdirectories
    """

    if files is None:
        my_name = os.path.splitext(os.path.basename(sys.argv[0]))[0]
        files = [my_name + ".csv", my_name + ".xlsx"] 

    for dir_name, _, file_names in os.walk(root_dir):
        for file_name in file_names:
            if (
                    (file_name.endswith(".csv") or file_name.endswith(".xlsx"))
                    and
                    (files is None or file_name in files or (
                        type(files) is str and fnmatch(file_name, files)
                    ))
                ):
                if verbose:
                    print "***", file_name
                yield os.path.join(dir_name, file_name)

